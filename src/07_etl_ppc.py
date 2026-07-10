"""
07_etl_ppc.py — Production Planning & Control business game (4 rounds).

Data-quality findings handled here (documented, not silently fixed):
  * ROUND_1 / ROUND_2 / ROUND_3.xlsx are content-identical copies of the
    Round 1 workbook — batch-level raw data for Rounds 2/3 was never saved.
    Round 2/3 summary KPIs are therefore taken from the accompanying report.
  * akhil_round_4_.xlsx actually contains the ROUND 3 measured data: its batch
    cycle times reproduce the report's Round-3 table exactly (7 of 8 batches).
  * Report's Round-3 batch 5 cycle time (58:59) mismatches the source data
    (59:49) — transposed digits in the report; source value is kept.
  * Round 1 contains broken #VALUE! formulas for station A3 — stored as NULL.
  * All times are Excel day-fractions — converted to minutes (x 1440).

Run:  python src/07_etl_ppc.py
"""

from pathlib import Path
import sqlite3

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
DB = ROOT / "outputs" / "battery_lab.db"

MIN = 1440.0  # Excel day fraction -> minutes


def num(cell):
    """Convert a cell to minutes: handles Excel day-fraction floats,
    datetime.time, datetime.timedelta, and broken '#VALUE!' strings (-> None)."""
    import datetime as dt
    v = cell.value
    if isinstance(v, dt.timedelta):
        return round(v.total_seconds() / 60.0, 3)
    if isinstance(v, dt.time):
        return round(v.hour * 60 + v.minute + v.second / 60.0
                     + v.microsecond / 6e7, 3)
    if isinstance(v, (int, float)):
        return round(v * MIN, 3)
    return None  # '#VALUE!' formula errors and empty cells


def parse_round1():
    ws = openpyxl.load_workbook(RAW / "ROUND_1.xlsx", data_only=True)["Round 1"]
    batches = [{"round": 1, "batch": b, "source": "measured (ROUND_1.xlsx)",
                "cycle_min": num(ws.cell(28 + b, 4)),       # D29..D36
                "waiting_min": num(ws.cell(28 + b, 10)),    # J29..J36
                "processing_min": num(ws.cell(41 + b, 4))}  # D42..D49
               for b in range(1, 9)]
    stations = []
    for i, name in enumerate(["Storage", "A1", "A2", "A3",
                              "Refining", "Quality Check", "Rework"]):
        stations.append({"round": 1, "workstation": name,
                         "operation_min": num(ws.cell(29 + i, 16)),   # P29..P35
                         "available_min": num(ws.cell(42 + i, 10)),   # J42..J48
                         "idle_min": num(ws.cell(42 + i, 16))})       # P42..P48
    return batches, stations


def parse_round3():
    """akhil_round_4_.xlsx — verified to hold the Round 3 measurements."""
    ws = openpyxl.load_workbook(RAW / "akhil_round_4_.xlsx", data_only=True)["Sheet1"]
    batches = [{"round": 3, "batch": b, "source": "measured (akhil_round_4_.xlsx)",
                "cycle_min": num(ws.cell(25 + b, 3)),       # C26..C33
                "waiting_min": num(ws.cell(25 + b, 11)),    # K26..K33
                "processing_min": num(ws.cell(44 + b, 3))}  # C45..C52
               for b in range(1, 9)]
    stations = []
    for i, name in enumerate(["Storage", "A1", "A2", "A3",
                              "Refining", "Quality Check", "Rework"]):
        stations.append({"round": 3, "workstation": name,
                         "operation_min": num(ws.cell(45 + i, 19)),   # S45..S51
                         "available_min": num(ws.cell(45 + i, 11)),   # K45..K51
                         "idle_min": num(ws.cell(25 + i + 1, 19))})   # S27..S33
    return batches, stations


# Round 2/3 summary KPIs and Round 4 (theoretical proposal): from the report,
# production_planning_presentation_.pdf — raw batch data was not preserved.
ROUND_SUMMARY = [
    (1, "Baseline (separate stations, central QC)", 125, 24, "measured"),
    (2, "A2 + A3 combined (Blender)", 75, 28, "report (raw file is a copy of Round 1)"),
    (3, "Decentralized quality control", 68, 30, "measured"),
    (4, "Lean optimized (std. work, pre-staging, 5S)", 58, 36, "theoretical proposal (report)"),
]

R4_PROPOSED = [  # report slides 9A / 9B (minutes)
    (1, 40, 10), (2, 45, 10), (3, 35, 11), (4, 34, 12),
    (5, 50, 8), (6, 53, 15), (7, 52, 17), (8, 54, 18),
]

SCHEMA = """
DROP TABLE IF EXISTS ppc_batch;
DROP TABLE IF EXISTS ppc_workstation;
DROP TABLE IF EXISTS ppc_round_summary;
CREATE TABLE ppc_batch (
    round INTEGER, batch INTEGER, cycle_min REAL, waiting_min REAL,
    processing_min REAL, source TEXT, PRIMARY KEY (round, batch));
CREATE TABLE ppc_workstation (
    round INTEGER, workstation TEXT, operation_min REAL,
    available_min REAL, idle_min REAL, PRIMARY KEY (round, workstation));
CREATE TABLE ppc_round_summary (
    round INTEGER PRIMARY KEY, label TEXT, horizon_min REAL,
    utilization_pct REAL, source TEXT);
"""


def main():
    b1, s1 = parse_round1()
    b3, s3 = parse_round3()
    b4 = [{"round": 4, "batch": b, "cycle_min": c, "waiting_min": w,
           "processing_min": None, "source": "theoretical proposal (report)"}
          for b, c, w in R4_PROPOSED]

    with sqlite3.connect(DB) as con:
        con.executescript(SCHEMA)
        con.executemany(
            """INSERT INTO ppc_batch VALUES
               (:round,:batch,:cycle_min,:waiting_min,:processing_min,:source)""",
            b1 + b3 + b4)
        con.executemany(
            """INSERT INTO ppc_workstation VALUES
               (:round,:workstation,:operation_min,:available_min,:idle_min)""",
            s1 + s3)
        con.executemany("INSERT INTO ppc_round_summary VALUES (?,?,?,?,?)",
                        ROUND_SUMMARY)
        for t in ["ppc_batch", "ppc_workstation", "ppc_round_summary"]:
            n = con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            print(f"{t:20s} {n:3d} rows")


if __name__ == "__main__":
    main()
