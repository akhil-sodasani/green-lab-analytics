"""
10_etl_pv.py — Photovoltaics lab: module IV curves from a Sinus IV tracer.

11 instrument exports (IV-Summary + IV-Raw + Celltype sheets each) covering:
  * an irradiance series at ~24 C (180 / 400 / 626 / 842 / 1053 W/m2)
  * a full-sun reference and two shading cases (1 cell, 2 cells parallel)
  * one lab measurement and two additional curves from an earlier session
plus a manually measured outdoor IV table transcribed from the lab report.

Data-quality / methodology notes handled here:
  * Actual irradiance and module temperature are encoded in the FILENAMES,
    not in the instrument summary (whose Temperature field is the 25 C
    reference). Both are parsed and stored per measurement.
  * The instrument's ETA assumes the STC reference irradiance (1000 W/m2).
    At e.g. 180 W/m2 that understates efficiency by 5x. The SQL layer
    recomputes efficiency with the ACTUAL irradiance.
  * Raw sweeps contain small negative currents near Voc (sensor offset) -
    kept as measured; the SQL KPI extraction handles them explicitly.

Run:  python src/10_etl_pv.py
"""

from pathlib import Path
import re
import sqlite3

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
DB = ROOT / "outputs" / "battery_lab.db"

FILES = {
    # file: (label, condition, irradiance W/m2, module temp C)
    "STC_23_99C_179_83Wm2.xlsx": ("180 W/m²", "irradiance_series", 179.83, 23.99),
    "STC_23_74C_400_42Wm2.xlsx": ("400 W/m²", "irradiance_series", 400.42, 23.74),
    "STC_23_6C_625_62Wm2.xlsx": ("626 W/m²", "irradiance_series", 625.62, 23.60),
    "STC_23_58C_841_81Wm2.xlsx": ("842 W/m²", "irradiance_series", 841.81, 23.58),
    "STC_25_1C_1052_6Wm2.xlsx": ("1053 W/m²", "irradiance_series", 1052.60, 25.10),
    "STC_36_02C_1051_24Wm2_100.xlsx": ("full sun (no shade)", "shading_ref", 1051.24, 36.02),
    "STC_30_71C_1057_14Wm2_1cell_shade.xlsx": ("1 cell shaded", "shading", 1057.14, 30.71),
    "STC_27_95C_1058_35Wm2_2C_PAR_shade.xlsx": ("2 cells (parallel) shaded", "shading", 1058.35, 27.95),
    "Lab_28_7C_731_14Wm2.xlsx": ("lab measurement", "lab", 731.14, 28.70),
    "IV-2_Curve_998__29_01C.xlsx": ("curve 2 (998 W/m²)", "earlier_session", 998.0, 29.01),
    "IV-4_Curve_516__32_82C.xlsx": ("curve 4 (516 W/m²)", "earlier_session", 516.0, 32.82),
}

SUMMARY_KEYS = ["Isc", "Voc", "FF", "Vmpp", "Impp", "Pmpp", "ETA", "Area"]

# Manual outdoor measurement transcribed from the lab report (resistance
# sweep 0-100 ohm, ~715 W/m2): R, T_C, U_V, I_A, G_Wm2
OUTDOOR_MANUAL = [
    (0, 26.42, 18.8, 0.86, 714), (10, 28.40, 18.6, 0.95, 714),
    (20, 30.64, 18.5, 1.06, 715), (30, 31.02, 18.3, 1.27, 715),
    (40, 31.23, 18.1, 1.40, 716), (50, 31.32, 17.8, 1.65, 717),
    (60, 31.54, 17.3, 2.04, 716), (70, 31.67, 16.4, 2.62, 716),
    (80, 31.86, 13.6, 3.45, 716), (90, 31.92, 6.4, 3.60, 716),
    (100, 32.04, 0.1, 3.65, 716),
]

SCHEMA = """
DROP TABLE IF EXISTS pv_measurement;
DROP TABLE IF EXISTS pv_iv_raw;
DROP TABLE IF EXISTS pv_outdoor_manual;
CREATE TABLE pv_measurement (
    dataset TEXT PRIMARY KEY, label TEXT, condition TEXT,
    irradiance_wm2 REAL, module_temp_c REAL, area_cm2 REAL,
    inst_isc_a REAL, inst_voc_v REAL, inst_ff_pct REAL,
    inst_vmpp_v REAL, inst_impp_a REAL, inst_pmpp_w REAL,
    inst_eta_pct REAL);
CREATE TABLE pv_iv_raw (
    dataset TEXT, voltage_v REAL, current_a REAL);
CREATE TABLE pv_outdoor_manual (
    resistance_ohm REAL, module_temp_c REAL, voltage_v REAL,
    current_a REAL, irradiance_wm2 REAL);
"""


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    summary = {}
    for row in wb["IV-Summary"].iter_rows(min_col=8, max_col=9):
        key, val = row[0].value, row[1].value
        if key in SUMMARY_KEYS and isinstance(val, (int, float)):
            summary[key] = round(val, 4)
    raw, seen = [], set()
    for row in wb["IV-Raw"].iter_rows(min_row=5, min_col=4, max_col=5):
        v, i = row[0].value, row[1].value
        if isinstance(v, (int, float)) and isinstance(i, (int, float)):
            raw.append((round(v, 5), round(i, 6)))
    return summary, raw


def main():
    with sqlite3.connect(DB) as con:
        con.executescript(SCHEMA)
        total = 0
        for fname, (label, cond, g, t) in FILES.items():
            summary, raw = parse_file(RAW / fname)
            ds = fname.replace(".xlsx", "")
            con.execute(
                "INSERT INTO pv_measurement VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (ds, label, cond, g, t, summary.get("Area"),
                 summary.get("Isc"), summary.get("Voc"), summary.get("FF"),
                 summary.get("Vmpp"), summary.get("Impp"),
                 summary.get("Pmpp"), summary.get("ETA")))
            con.executemany(
                "INSERT INTO pv_iv_raw VALUES (?,?,?)",
                [(ds, v, i) for v, i in raw])
            total += len(raw)
        con.executemany("INSERT INTO pv_outdoor_manual VALUES (?,?,?,?,?)",
                        OUTDOOR_MANUAL)
        print(f"pv_measurement: {len(FILES)} datasets · pv_iv_raw: {total} points "
              f"· pv_outdoor_manual: {len(OUTDOOR_MANUAL)} rows")


if __name__ == "__main__":
    main()
