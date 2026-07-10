"""
04_etl_re05.py — Extract & clean the RE05 hydrogen-technologies lab data plus the
capacitor/regulator transient recordings, and load into outputs/battery_lab.db.

Sources (data/raw/):
    RE_05_grp_4.xlsx          spatially arranged multi-experiment sheet:
                              electrolyzer runs (2 loads), electrolyzer I-U,
                              PEM cell U-I/P, 13-cell PEM stack run, SOFC U-I,
                              SOFC constant-load run, burner masses (German commas)
    Mappe1/2/3.xlsx           capacitor discharge transients (10 ms sampling)
    deep_dischaRGE_prot.xlsx  two slow charge/load recordings (side-by-side blocks)
    Regulators_.xlsx          five regulator topologies side by side

Run:  python src/04_etl_re05.py
"""

from pathlib import Path
import sqlite3

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
DB = ROOT / "outputs" / "battery_lab.db"


def de_float(text):
    """Parse German decimal commas out of strings like 'cell weight=89,9205 g'."""
    digits = "".join(ch for ch in str(text) if ch.isdigit() or ch in ",.")
    return float(digits.replace(",", "."))


def parse_re05(path):
    ws = openpyxl.load_workbook(path, data_only=True)["Tabelle1"]
    v = lambda coord: ws[coord].value

    # -- electrolyzer runs: two loads, current every 30 s, gas volume at end ----
    runs = []
    for r in range(3, 13):
        runs.append({"load_ohm": 5.0, "time_s": v(f"B{r}"),
                     "current_ma": v(f"A{r}"), "voltage_v": v(f"D{r}"),
                     "h2_volume_ml": v(f"E{r}")})
        runs.append({"load_ohm": 3.0, "time_s": v(f"H{r}"),
                     "current_ma": v(f"G{r}"), "voltage_v": v(f"J{r}"),
                     "h2_volume_ml": v(f"K{r}")})
    electrolyzer_run = pd.DataFrame(runs).dropna(subset=["time_s"])
    # duplicate volume reading at t=180 for the 3-ohm run (K9) is an
    # intermediate reading; end-of-run volume is the one at t=270.

    # -- electrolyzer I-U characteristic ---------------------------------------
    electrolyzer_iu = pd.DataFrame(
        [{"current_ma": v(f"A{r}"), "voltage_v": v(f"B{r}"),
          "resistance_ohm": v(f"C{r}")} for r in range(17, 30)])

    # -- PEM single-cell U-I / power (voltage recorded in mV) ------------------
    pem_cell_iu = pd.DataFrame(
        [{"current_ma": v(f"F{r}"), "voltage_mv": v(f"G{r}"),
          "resistance_ohm": v(f"H{r}"),
          "setup": 2 if r >= 21 else 1}          # 'new setup' noted at 10 ohm row
         for r in range(17, 28)]).dropna(subset=["current_ma"])

    # -- PEM 13-cell stack, constant 1.78 A ------------------------------------
    pem_stack_run = pd.DataFrame(
        [{"time_s": v(f"B{r}"), "current_a": v(f"A{r}"), "power_w": v(f"C{r}")}
         for r in range(36, 47)])
    # power at t=0 was not recorded in the protocol (missing value kept as NULL)

    # -- SOFC U-I characteristic (mV / mA) --------------------------------------
    sofc_iu = pd.DataFrame(
        [{"voltage_mv": v(f"G{r}"), "current_ma": v(f"H{r}"),
          "resistance_ohm": v(f"I{r}")} for r in range(43, 53)])

    # -- SOFC constant-load run at 2 ohm ----------------------------------------
    sofc_run = pd.DataFrame(
        [{"time_s": v(f"A{r}"), "voltage_mv": v(f"B{r}"),
          "current_ma": v(f"C{r}"), "resistance_ohm": v(f"D{r}")}
         for r in range(57, 64)])

    # -- burner / cell masses (German decimal commas in free-text cells) --------
    burner_mass = pd.DataFrame([
        {"label": "pem_block_weighing", "before_g": de_float(v("H34")),
         "after_g": de_float(v("H35")), "note": "labelled 'cell weight', 'end at 5 mins'"},
        {"label": "sofc_burner", "before_g": de_float(v("H59")),
         "after_g": de_float(v("H60")), "note": "adjacent to SOFC 2-ohm run"},
    ])
    return (electrolyzer_run, electrolyzer_iu, pem_cell_iu,
            pem_stack_run, sofc_iu, sofc_run, burner_mass)


def parse_cap_discharges():
    frames = []
    for run, fname in enumerate(["Mappe1.xlsx", "Mappe_2.xlsx", "Mappe3.xlsx"], 1):
        df = pd.read_excel(RAW / fname)
        frames.append(pd.DataFrame({
            "run": run, "time_s": df["t / s"],
            "current_a": df["I_A1 / A"], "voltage_v": df["U_B1 / V"]}))
    return pd.concat(frames, ignore_index=True)


def parse_regulators():
    """Five topology blocks side by side; auto-detect via 't / s' header cells."""
    ws = openpyxl.load_workbook(RAW / "Regulators_.xlsx", data_only=True)["Tabelle1"]
    labels = ["Series Regulator", "Shunt Regulator", "PWM",
              "Shunt Regulator + Load", "MPP + Shunt"]
    starts = [c.column for c in ws[2] if c.value == "t / s"]
    frames = []
    for label, col in zip(labels, starts):
        rows = []
        for r in range(3, ws.max_row + 1):
            t = ws.cell(r, col).value
            if t is None:
                break
            rows.append({"topology": label, "time_s": t,
                         "i_in_a": ws.cell(r, col + 1).value,
                         "u_in_v": ws.cell(r, col + 2).value,
                         "i_cap_a": ws.cell(r, col + 3).value,
                         "u_cap_v": ws.cell(r, col + 4).value})
        frames.append(pd.DataFrame(rows))
    return pd.concat(frames, ignore_index=True)


SCHEMA = """
DROP TABLE IF EXISTS electrolyzer_run;
DROP TABLE IF EXISTS electrolyzer_iu;
DROP TABLE IF EXISTS pem_cell_iu;
DROP TABLE IF EXISTS pem_stack_run;
DROP TABLE IF EXISTS sofc_iu;
DROP TABLE IF EXISTS sofc_run;
DROP TABLE IF EXISTS burner_mass;
DROP TABLE IF EXISTS cap_discharge;
DROP TABLE IF EXISTS regulator_run;

CREATE TABLE electrolyzer_run (
    load_ohm REAL, time_s REAL, current_ma REAL,
    voltage_v REAL, h2_volume_ml REAL,
    PRIMARY KEY (load_ohm, time_s));
CREATE TABLE electrolyzer_iu (
    current_ma REAL, voltage_v REAL, resistance_ohm REAL);
CREATE TABLE pem_cell_iu (
    current_ma REAL, voltage_mv REAL, resistance_ohm REAL, setup INTEGER);
CREATE TABLE pem_stack_run (
    time_s REAL PRIMARY KEY, current_a REAL, power_w REAL);
CREATE TABLE sofc_iu (
    voltage_mv REAL, current_ma REAL, resistance_ohm REAL);
CREATE TABLE sofc_run (
    time_s REAL PRIMARY KEY, voltage_mv REAL, current_ma REAL, resistance_ohm REAL);
CREATE TABLE burner_mass (
    label TEXT PRIMARY KEY, before_g REAL, after_g REAL, note TEXT);
CREATE TABLE cap_discharge (
    run INTEGER, time_s REAL, current_a REAL, voltage_v REAL,
    PRIMARY KEY (run, time_s));
CREATE TABLE regulator_run (
    topology TEXT, time_s REAL, i_in_a REAL, u_in_v REAL,
    i_cap_a REAL, u_cap_v REAL,
    PRIMARY KEY (topology, time_s));
"""


def main():
    tables = parse_re05(RAW / "RE_05_grp_4.xlsx")
    names = ["electrolyzer_run", "electrolyzer_iu", "pem_cell_iu",
             "pem_stack_run", "sofc_iu", "sofc_run", "burner_mass"]
    cap = parse_cap_discharges()
    reg = parse_regulators()

    with sqlite3.connect(DB) as con:
        con.executescript(SCHEMA)
        for name, df in zip(names, tables):
            df.to_sql(name, con, if_exists="append", index=False)
        cap.to_sql("cap_discharge", con, if_exists="append", index=False)
        reg.to_sql("regulator_run", con, if_exists="append", index=False)
        counts = {name: con.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
                  for name in names + ["cap_discharge", "regulator_run"]}
    for name, n in counts.items():
        print(f"{name:18s} {n:5d} rows")


if __name__ == "__main__":
    main()
