"""
01_etl.py — Extract & clean the raw lab workbook into tidy tables, load into SQLite.

The source file is a single messy sheet containing four experiment blocks side by
side, German decimal commas, a typo'd broken formula, and merged header rows.
This script turns it into 4 normalized tables:

    faraday_experiment   (1 row  — copper electrolysis parameters & masses)
    battery_soc          (4 rows — open-circuit voltage / state of charge)
    ui_curve             (36 rows — U-I characteristic, 3 chemistries x 12 loads)
    cycle_test           (22 rows — NiMH discharge/charge voltage time series)

Run:  python src/01_etl.py
"""

from pathlib import Path
import sqlite3

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "Electrochemical_storages.xlsx"
DB = ROOT / "outputs" / "battery_lab.db"


def to_float(value):
    """Handle German decimal commas ('4,2' -> 4.2) and stray strings."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip().replace(",", "."))


def extract(ws):
    # ---- Experiment 1: Faraday constant (copper deposition) -----------------
    faraday = pd.DataFrame([{
        "electrode1_before_g": to_float(ws["C4"].value),
        "electrode2_before_g": to_float(ws["C5"].value),
        "electrode1_after_g": to_float(ws["C11"].value),
        "electrode2_after_g": to_float(ws["C12"].value),
        "current_a": to_float(ws["C6"].value),
        "time_min": to_float(ws["C7"].value),
    }])

    # ---- Experiment 2: open-circuit voltage / state of charge ---------------
    soc_rows = []
    names = {"Li polymer": "LiPo", "pb battery": "Lead-acid",
             "NiZn": "NiZn", "NiMH": "NiMH"}
    for r in range(3, 7):
        soc_rows.append({
            "battery": names[ws[f"E{r}"].value.strip()],
            "v_max": to_float(ws[f"F{r}"].value),
            "ocv_measured_v": to_float(ws[f"G{r}"].value),
            "mean_working_voltage_v": to_float(ws[f"H{r}"].value),
            "weight_g": to_float(ws[f"J{r}"].value),
            "soc": to_float(ws[f"L{r}"].value),
            "capacity_max_mah": to_float(ws[f"M{r}"].value),
            "capacity_current_mah": to_float(ws[f"N{r}"].value),
        })
    battery_soc = pd.DataFrame(soc_rows)
    # NOTE: the source column "current E density" contained a broken formula
    # (=M2*G2/I2 pointed at header cells) for LiPo. Energy density is
    # deliberately NOT extracted here — it is recomputed cleanly in SQL.

    # ---- Experiment 3: U-I characteristic, 3 chemistries --------------------
    ui_rows = []
    cols = {"Li-Ion": ("R", "S"), "NiMH": ("T", "U"), "Lead-acid": ("V", "W")}
    for r in range(3, 15):
        resistance = to_float(ws[f"Q{r}"].value)
        for battery, (u_col, i_col) in cols.items():
            ui_rows.append({
                "battery": battery,
                "load_resistance_ohm": resistance,
                "voltage_v": to_float(ws[f"{u_col}{r}"].value),
                "current_ma": to_float(ws[f"{i_col}{r}"].value),
            })
    ui_curve = pd.DataFrame(ui_rows)

    # ---- Experiment 4: NiMH discharge / charge cycle (target 230 mA) --------
    cycle_rows = []
    for r in range(19, 30):
        cycle_rows.append({"phase": "discharge",
                           "time_s": to_float(ws[f"Q{r}"].value),
                           "voltage_v": to_float(ws[f"R{r}"].value)})
        cycle_rows.append({"phase": "charge",
                           "time_s": to_float(ws[f"U{r}"].value),
                           "voltage_v": to_float(ws[f"V{r}"].value)})
    cycle_test = pd.DataFrame(cycle_rows)
    cycle_test["current_ma"] = 230.0

    return faraday, battery_soc, ui_curve, cycle_test


def main():
    ws = openpyxl.load_workbook(RAW, data_only=True)["Tabelle1"]
    faraday, battery_soc, ui_curve, cycle_test = extract(ws)

    DB.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB) as con:
        schema = (ROOT / "sql" / "schema.sql").read_text()
        con.executescript(schema)
        faraday.to_sql("faraday_experiment", con, if_exists="append", index=False)
        battery_soc.to_sql("battery_soc", con, if_exists="append", index=False)
        ui_curve.to_sql("ui_curve", con, if_exists="append", index=False)
        cycle_test.to_sql("cycle_test", con, if_exists="append", index=False)

    print(f"Loaded {len(faraday)} + {len(battery_soc)} + {len(ui_curve)} + "
          f"{len(cycle_test)} rows into {DB.name}")


if __name__ == "__main__":
    main()
